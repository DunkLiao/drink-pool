import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from models import now


def export_orders_to_excel(session):
    wb = Workbook()
    ws = wb.active
    ws.title = '團購訂單'

    # Header styles
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font_white = Font(bold=True, size=12, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    center_align = Alignment(horizontal='center', vertical='center')
    wrap_align = Alignment(vertical='center', wrap_text=True)

    # Title row
    ws.merge_cells('A1:I1')
    title_cell = ws['A1']
    title_cell.value = f'【{session.title}】團購訂單彙整'
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 36

    # Info row
    ws.merge_cells('A2:I2')
    info_cell = ws['A2']
    time_range = f'{session.start_time.strftime("%Y/%m/%d %H:%M")} ~ {session.end_time.strftime("%Y/%m/%d %H:%M")}'
    info_cell.value = f'團購期間：{time_range}　　匯出時間：{now().strftime("%Y/%m/%d %H:%M:%S")}'
    info_cell.font = Font(size=10, color='666666')
    info_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 24

    # Headers
    headers = ['編號', '姓名', '科別', '飲料品項', '單價', '甜度', '冰塊', '加料', '備註']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    ws.row_dimensions[4].height = 28

    # Data rows
    for row_idx, order in enumerate(session.orders, 1):
        addons_text = '、'.join(a.addon_name for a in sorted(order.addons, key=lambda x: x.sort_order))
        row_data = [
            row_idx,
            order.name,
            order.department.name if order.department else '',
            order.drink_item,
            order.drink_price if order.drink_price is not None else '',
            order.sweetness,
            order.ice,
            addons_text,
            order.notes or '',
        ]
        excel_row = row_idx + 4
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx in (1, 5, 6, 7):
                cell.alignment = center_align
            else:
                cell.alignment = wrap_align
        ws.row_dimensions[excel_row].height = 22

    # Summary row
    summary_row = len(session.orders) + 5
    ws.merge_cells(f'A{summary_row}:I{summary_row}')
    summary_cell = ws.cell(row=summary_row, column=1)
    summary_cell.value = f'共 {len(session.orders)} 筆訂單'
    summary_cell.font = Font(bold=True, size=11)
    summary_cell.alignment = Alignment(horizontal='right', vertical='center')
    summary_cell.border = thin_border

    # Column widths
    col_widths = [8, 16, 16, 30, 10, 12, 12, 30, 30]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
